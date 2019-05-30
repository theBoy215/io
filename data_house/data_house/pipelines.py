# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html

import openpyxl


class DataHousePipeline(object):
    def __init__(self):
        # self.work = openpyxl.Workbook()
        self.work = openpyxl.load_workbook('安居客廊坊三河房源信息.xlsx')
        self.sheet1 = self.work.create_sheet('二手房', index=0)
        self.sheet1['A1'] = '售价'
        self.sheet1['B1'] = '单价'
        self.sheet1['C1'] = '面积'
        self.sheet1['D1'] = '经度'
        self.sheet1['E1'] = '纬度'
        self.count = 2

    def process_item(self, item, spider):
        self.sheet1['A%d' % self.count] = item['price']
        self.sheet1['B%d' % self.count] = item['sprice']
        self.sheet1['C%d' % self.count] = item['meter']
        self.sheet1['D%d' % self.count] = item['pos_x']
        self.sheet1['E%d' % self.count] = item['pos_y']
        self.count += 1
        return item

    def close_spider(self, spider):
        self.work.save('安居客廊坊三河房源信息.xlsx')
