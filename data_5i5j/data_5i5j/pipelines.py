# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import openpyxl


class Data5I5JPipeline(object):
    def __init__(self):
        self.work = openpyxl.Workbook()
        self.sheet1 = self.work.create_sheet('整租', index=0)
        self.sheet1['A1'] = '标题'
        self.sheet1['B1'] = '坐标'
        self.sheet1['C1'] = '图片地址'
        self.sheet1['D1'] = '售价(万)'
        self.sheet1['E1'] = '单价(万/m²)'
        self.sheet1['F1'] = '标签'
        self.sheet1['G1'] = '户型'
        self.sheet1['H1'] = '面积'
        self.sheet1['I1'] = '朝向'
        self.sheet1['J1'] = '装修'
        self.sheet1['K1'] = '建筑类型'
        self.sheet1['L1'] = '楼层'
        self.sheet1['M1'] = '年代'
        self.sheet1['N1'] = '商圈'
        self.sheet1['O1'] = '小区'
        self.count = 2

    def process_item(self, item, spider):
        self.sheet1['A%d' % self.count] = item['title']
        self.sheet1['B%d' % self.count] = item['location']
        self.sheet1['C%d' % self.count] = item['img_url']
        self.sheet1['D%d' % self.count] = item['price']
        self.sheet1['E%d' % self.count] = item['unitprice']
        self.sheet1['F%d' % self.count] = item['tags']
        self.sheet1['G%d' % self.count] = item['layout']
        self.sheet1['H%d' % self.count] = item['area']
        self.sheet1['I%d' % self.count] = item['heading']
        self.sheet1['J%d' % self.count] = item['decoration']
        self.sheet1['K%d' % self.count] = item['buildingtype']
        self.sheet1['L%d' % self.count] = item['floor']
        self.sheet1['M%d' % self.count] = item['buildyear']
        self.sheet1['N%d' % self.count] = item['sq']
        self.sheet1['O%d' % self.count] = item['communityname']
        self.count += 1
        return item

    def close_spider(self, spider):
        self.work.save('我爱我家昌平区租房房源信息.xlsx')


class DataZ5I5JPipeline(object):
    def __init__(self):
        self.work = openpyxl.Workbook()
        self.sheet1 = self.work.create_sheet('整租', index=0)
        self.sheet1['A1'] = '标题'
        self.sheet1['B1'] = '副标题'
        self.sheet1['C1'] = '租金 (元/月)'
        self.sheet1['D1'] = '户型'
        self.sheet1['E1'] = '面积(平米)'
        self.sheet1['F1'] = '支付方式'
        self.sheet1['G1'] = '小区'
        self.sheet1['H1'] = '楼层'
        self.sheet1['I1'] = '朝向'
        self.sheet1['J1'] = '装修'
        self.sheet1['K1'] = '地铁'
        self.count = 2

    def process_item(self, item, spider):
        self.sheet1['A%d' % self.count] = item['title']
        self.sheet1['B%d' % self.count] = item['sub_t']
        self.sheet1['C%d' % self.count] = item['price']
        self.sheet1['D%d' % self.count] = item['buildingtype']
        self.sheet1['E%d' % self.count] = item['area']
        self.sheet1['F%d' % self.count] = item['pays']
        self.sheet1['G%d' % self.count] = item['communityname']
        self.sheet1['H%d' % self.count] = item['floor']
        self.sheet1['I%d' % self.count] = item['heading']
        self.sheet1['J%d' % self.count] = item['decoration']
        self.sheet1['K%d' % self.count] = item['sub']
        self.count += 1
        return item

    def close_spider(self, spider):
        self.work.save(r'C:\Users\Administrator\PycharmProjects\spiders\租房信息\data_5i5j\我爱我家租房房源信息\我爱我家昌平区整租房源信息1.xlsx')


class DataLianJiaPipeline(object):
    def __init__(self):
        self.work = openpyxl.load_workbook('延庆.xlsx')
        self.sheet1 = self.work.create_sheet(r'二手房', index=0)
        self.sheet1['A1'] = '标题'
        self.sheet1['B1'] = '售价'
        self.sheet1['C1'] = '单价'
        self.sheet1['D1'] = '室'
        self.sheet1['E1'] = '厅'
        self.sheet1['F1'] = '面积'
        self.sheet1['G1'] = '朝向'
        self.sheet1['H1'] = '楼层'
        self.sheet1['I1'] = '总楼层'
        self.sheet1['J1'] = '电梯'
        self.sheet1['K1'] = '小区名称'
        self.sheet1['L1'] = '地铁距离'
        self.sheet1['M1'] = '经度'
        self.sheet1['N1'] = '纬度'
        self.count = 2

    def process_item(self, item, spider):
        self.sheet1['A%d' % self.count] = item['title']
        self.sheet1['B%d' % self.count] = item['price']
        self.sheet1['C%d' % self.count] = item['unitprice']
        self.sheet1['D%d' % self.count] = item['s']
        self.sheet1['E%d' % self.count] = item['t']
        self.sheet1['F%d' % self.count] = item['area']
        self.sheet1['G%d' % self.count] = item['threading']
        self.sheet1['H%d' % self.count] = item['floor']
        self.sheet1['I%d' % self.count] = item['floors']
        self.sheet1['J%d' % self.count] = item['lift']
        self.sheet1['K%d' % self.count] = item['areaName']
        self.sheet1['L%d' % self.count] = item['sub_met']
        self.sheet1['M%d' % self.count] = item['coord_x']
        self.sheet1['N%d' % self.count] = item['coord_y']
        self.count += 1
        return item

    def close_spider(self, spider):
        self.work.save('延庆.xlsx')


class DataZLianJiaPipeline(object):
    def __init__(self):
        # self.work = openpyxl.Workbook()
        self.work = openpyxl.load_workbook('链家房山租房房源信息.xlsx')
        # self.sheet1 = self.work.create_sheet('润枫欣尚', index=0)
        self.sheet1 = self.work.create_sheet('整租', index=0)
        self.sheet1['A1'] = '标题'
        self.sheet1['B1'] = '标签'
        self.sheet1['C1'] = '租金'
        self.sheet1['D1'] = '室'
        self.sheet1['E1'] = '厅'
        self.sheet1['F1'] = '卫'
        self.sheet1['G1'] = '面积'
        self.sheet1['H1'] = '朝向'
        self.sheet1['I1'] = '楼层'
        self.sheet1['J1'] = '总楼层'
        self.sheet1['K1'] = '小区'
        self.sheet1['L1'] = '有无电梯'
        self.sheet1['M1'] = '坐标'
        self.sheet1['N1'] = '地铁距离'
        self.sheet1['O1'] = '医院'
        self.sheet1['P1'] = '公交站'
        self.sheet1['Q1'] = '超市'
        self.sheet1['R1'] = '付款方式'
        self.sheet1['S1'] = '押金'
        self.count = 2

    def process_item(self, item, spider):
        self.sheet1['A%d' % self.count] = item['title']
        self.sheet1['B%d' % self.count] = item['tag']
        self.sheet1['C%d' % self.count] = item['zprice']
        self.sheet1['D%d' % self.count] = item['type_s']
        self.sheet1['E%d' % self.count] = item['type_t']
        self.sheet1['F%d' % self.count] = item['type_w']
        self.sheet1['G%d' % self.count] = item['area']
        self.sheet1['H%d' % self.count] = item['threading']
        self.sheet1['I%d' % self.count] = item['floor']
        self.sheet1['J%d' % self.count] = item['floors']
        self.sheet1['K%d' % self.count] = item['areaName']
        self.sheet1['L%d' % self.count] = item['lift']
        self.sheet1['M%d' % self.count] = item['coord']
        self.sheet1['N%d' % self.count] = item['sub_met']
        self.sheet1['O%d' % self.count] = item['yiy']
        self.sheet1['P%d' % self.count] = item['bus']
        self.sheet1['Q%d' % self.count] = item['shop']
        self.sheet1['R%d' % self.count] = item['pry_type']
        self.sheet1['S%d' % self.count] = item['y_price']
        self.count += 1
        return item

    def close_spider(self, spider):
        self.work.save('链家房山租房房源信息.xlsx')


class DataZiRuPipeline(object):
    def __init__(self):
        self.work = openpyxl.load_workbook('自如亦庄开发区租房房源信息.xlsx')
        self.sheet1 = self.work.create_sheet('合租', index=1)
        self.sheet1['A1'] = '标题'
        self.sheet1['B1'] = '租金'
        self.sheet1['C1'] = '面积'
        self.sheet1['D1'] = '朝向'
        self.sheet1['E1'] = '室'
        self.sheet1['F1'] = '厅'
        self.sheet1['G1'] = '楼层'
        self.sheet1['H1'] = '总楼层'
        self.sheet1['I1'] = '地铁距离'
        self.sheet1['J1'] = '经度'
        self.sheet1['K1'] = '纬度'
        # self.sheet1['L1'] = '望京'
        # self.sheet1['M1'] = '上地'
        # self.sheet1['N1'] = '中关村'
        # self.sheet1['O1'] = '西二旗'
        # self.sheet1['P1'] = '国贸'
        # self.sheet1['Q1'] = '西单'
        self.count = 2

    def process_item(self, item, spider):
        self.sheet1['A%d' % self.count] = item['title']
        self.sheet1['B%d' % self.count] = item['price']
        self.sheet1['C%d' % self.count] = item['meter']
        self.sheet1['D%d' % self.count] = item['threading']
        self.sheet1['E%d' % self.count] = item['type_s']
        self.sheet1['F%d' % self.count] = item['type_t']
        self.sheet1['G%d' % self.count] = item['floor']
        self.sheet1['H%d' % self.count] = item['floors']
        self.sheet1['I%d' % self.count] = item['sub']
        self.sheet1['J%d' % self.count] = item['position_x']
        self.sheet1['K%d' % self.count] = item['position_y']
        # self.sheet1['L%d' % self.count] = item['wj_time']
        # self.sheet1['M%d' % self.count] = item['sd_time']
        # self.sheet1['N%d' % self.count] = item['zgc_time']
        # self.sheet1['O%d' % self.count] = item['xeq_time']
        # self.sheet1['P%d' % self.count] = item['gm_time']
        # self.sheet1['Q%d' % self.count] = item['xd_time']
        self.count += 1
        return item

    def close_spider(self, spider):
        self.work.save('自如亦庄开发区租房房源信息.xlsx')


class DataCjLianJiaPipeline(object):
    def __init__(self):
        self.work = openpyxl.load_workbook('西城.xlsx')
        self.sheet1 = self.work.create_sheet('二手房', index=0)
        self.sheet1['A1'] = '标题'
        self.sheet1['B1'] = '挂牌时间'
        self.sheet1['C1'] = '成交时间'
        self.sheet1['D1'] = '挂牌价格'
        self.sheet1['E1'] = '成交价格'
        self.sheet1['F1'] = '单价'
        self.sheet1['G1'] = '成交周期'
        self.sheet1['H1'] = '调价'
        self.sheet1['I1'] = '带看次数'
        self.sheet1['J1'] = '关注人数'
        self.sheet1['K1'] = '浏览次数'
        self.sheet1['L1'] = '室'
        self.sheet1['M1'] = '厅'
        self.sheet1['N1'] = '厨'
        self.sheet1['O1'] = '卫'
        self.sheet1['P1'] = '楼层数'
        self.sheet1['Q1'] = '总楼层'
        self.sheet1['R1'] = '面积'
        self.sheet1['S1'] = '房屋朝向'
        self.sheet1['T1'] = '建造年限'
        self.sheet1['U1'] = '装修情况'
        self.sheet1['V1'] = '有无电梯'
        self.sheet1['W1'] = '经度'
        self.sheet1['X1'] = '纬度'
        self.sheet1['Y1'] = '小区'
        self.count = 2

    def process_item(self, item, spider):
        self.sheet1['A%d' % self.count] = item['title']
        self.sheet1['B%d' % self.count] = item['gptime']
        self.sheet1['C%d' % self.count] = item['cjtime']
        self.sheet1['D%d' % self.count] = item['gprice']
        self.sheet1['E%d' % self.count] = item['cjprice']
        self.sheet1['F%d' % self.count] = item['dprice']
        self.sheet1['G%d' % self.count] = item['zq']
        self.sheet1['H%d' % self.count] = item['tj']
        self.sheet1['I%d' % self.count] = item['watch']
        self.sheet1['J%d' % self.count] = item['gz']
        self.sheet1['K%d' % self.count] = item['ll']
        self.sheet1['L%d' % self.count] = item['type_s']
        self.sheet1['M%d' % self.count] = item['type_t']
        self.sheet1['N%d' % self.count] = item['type_c']
        self.sheet1['O%d' % self.count] = item['type_w']
        self.sheet1['P%d' % self.count] = item['floor']
        self.sheet1['Q%d' % self.count] = item['floors']
        self.sheet1['R%d' % self.count] = item['area']
        self.sheet1['S%d' % self.count] = item['threading']
        self.sheet1['T%d' % self.count] = item['year']
        self.sheet1['U%d' % self.count] = item['decoration']
        self.sheet1['V%d' % self.count] = item['lift']
        self.sheet1['W%d' % self.count] = item['coord_x']
        self.sheet1['X%d' % self.count] = item['coord_y']
        self.sheet1['Y%d' % self.count] = item['name']
        self.count += 1
        return item

    def close_spider(self, spider):
        self.work.save('西城.xlsx')
